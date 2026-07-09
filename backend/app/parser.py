import openpyxl
from datetime import datetime, date, timedelta
from typing import Optional


def _combine(day: date, t) -> Optional[datetime]:
    if t is None:
        return None
    if isinstance(t, datetime):
        return datetime.combine(day, t.time())
    return None


def _zone_short(zone_id: str) -> str:
    if not zone_id:
        return ""
    parts = [p.strip() for p in zone_id.split("/") if p.strip()]
    return parts[-1] if parts else zone_id


def parse_excel(file_path: str, settings: dict) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # ── GlobalParameter ──────────────────────────────────────────────────
    ws = wb["GlobalParameter"]
    rows = list(ws.iter_rows(values_only=True))
    raw_date = rows[1][0] if len(rows) > 1 else None
    if isinstance(raw_date, datetime):
        day = raw_date.date()
    elif isinstance(raw_date, date):
        day = raw_date
    else:
        wb.close()
        raise ValueError(f"Cannot read date from GlobalParameter: {raw_date}")

    # ── Agents ───────────────────────────────────────────────────────────
    ws = wb["Agent"]
    agents = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        # firstName, id, lastName, defaultShift.id, profile.id
        if row[1]:
            agents[row[1]] = {
                "id": row[1],
                "first_name": row[0] or "",
                "last_name": row[2] or "",
            }

    # ── PlannedAssignment ─────────────────────────────────────────────────
    ws = wb["PlannedAssignment"]
    # assignmentType, endTime, frozen, id, isTimeConstrained,
    # maintenancePassIdx, notStartable, startTime, agent.id, zone.id
    planned = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        mission_id = row[3]
        if not mission_id:
            continue
        planned_start = _combine(day, row[7])
        planned_end = _combine(day, row[1])
        agent_id = row[8]
        zone_id = row[9]

        if planned_start is None or planned_end is None:
            continue

        # Handle missions crossing midnight
        if planned_end < planned_start:
            planned_end += timedelta(days=1)

        planned_duration_min = (planned_end - planned_start).total_seconds() / 60
        if planned_duration_min <= 0:
            continue

        planned[mission_id] = {
            "agent_id": agent_id,
            "zone_id": zone_id,
            "zone_short": _zone_short(zone_id),
            "planned_start": planned_start,
            "planned_end": planned_end,
            "planned_duration_min": planned_duration_min,
        }

    # ── AssignmentExecutionData ───────────────────────────────────────────
    ws = wb["AssignmentExecutionData"]
    # message, timestamp, type, plannedAssignment.id
    events = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        mission_id = row[3]
        event_type = row[2]
        ts = row[1]
        if not mission_id or not event_type or not ts:
            continue
        if mission_id not in events:
            events[mission_id] = {}
        # Keep first START and last END (in case of duplicates)
        if event_type == "START" and "START" not in events[mission_id]:
            events[mission_id]["START"] = ts
        elif event_type == "END":
            events[mission_id]["END"] = ts

    wb.close()

    # ── Batch detection ───────────────────────────────────────────────────
    # Strategy 1: cluster END timestamps within batch_window seconds
    # Strategy 2: END timestamp exactly matches planned_end (job set it to planned time)
    batch_window = int(settings.get("batch_window_seconds", 30))
    batch_min_size = int(settings.get("batch_min_size", 3))

    end_ts_list = sorted(
        [(evs["END"], mid) for mid, evs in events.items() if "END" in evs],
        key=lambda x: x[0],
    )

    batch_by_cluster = set()
    i = 0
    while i < len(end_ts_list):
        j = i + 1
        while j < len(end_ts_list):
            delta = (end_ts_list[j][0] - end_ts_list[i][0]).total_seconds()
            if delta <= batch_window:
                j += 1
            else:
                break
        if j - i >= batch_min_size:
            for k in range(i, j):
                batch_by_cluster.add(end_ts_list[k][1])
        i += 1

    # Strategy 2: START and END both exactly match planned times (job sets them to planned values)
    # This catches the morning job that validates missions at their exact planned times.
    batch_by_exact = set()
    for mid, evs in events.items():
        if mid not in planned:
            continue
        has_start = "START" in evs
        has_end = "END" in evs
        if not (has_start and has_end):
            continue
        delta_start = abs((evs["START"] - planned[mid]["planned_start"]).total_seconds())
        delta_end = abs((evs["END"] - planned[mid]["planned_end"]).total_seconds())
        if delta_start <= 2 and delta_end <= 2:
            batch_by_exact.add(mid)
    # Only apply if enough missions match (avoid false-positives for genuinely on-time agents)
    if len(batch_by_exact) < batch_min_size:
        batch_by_exact = set()

    batch_missions = batch_by_cluster | batch_by_exact

    # ── Build mission records ─────────────────────────────────────────────
    threshold_low = float(settings.get("threshold_low", 0.10))
    threshold_high = float(settings.get("threshold_high", 3.00))

    missions = []
    for mission_id, plan in planned.items():
        evs = events.get(mission_id, {})
        actual_start = evs.get("START")
        actual_end = evs.get("END")
        is_batch = mission_id in batch_missions

        actual_duration_min = None
        ratio = None
        anomaly_type = None
        status = "not_started"

        if actual_start:
            status = "in_progress"
        if actual_start and actual_end:
            status = "completed"
            actual_duration_min = (actual_end - actual_start).total_seconds() / 60
            if plan["planned_duration_min"] > 0:
                ratio = actual_duration_min / plan["planned_duration_min"]
                if not is_batch:
                    if ratio < threshold_low:
                        anomaly_type = "too_quick"
                    elif ratio > threshold_high:
                        anomaly_type = "too_long"

        missions.append(
            {
                "original_id": mission_id,
                "agent_id": plan["agent_id"],
                "zone_id": plan["zone_id"],
                "zone_short": plan["zone_short"],
                "planned_start": plan["planned_start"],
                "planned_end": plan["planned_end"],
                "planned_duration_min": plan["planned_duration_min"],
                "actual_start": actual_start,
                "actual_end": actual_end,
                "actual_duration_min": actual_duration_min,
                "ratio": ratio,
                "is_batch": is_batch,
                "anomaly_type": anomaly_type,
                "status": status,
            }
        )

    return {
        "day_date": day.isoformat(),
        "agents": list(agents.values()),
        "missions": missions,
    }
